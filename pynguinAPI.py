import re
import subprocess
import shlex
import os
import utils
from find_function_parameters import *
from openpyxl import load_workbook

TARGET_SCRIPT = "./pynguinAPI.sh" if utils.is_unix() else "./pynguinAPI.ps"
# INPUT_DIR and OUTPUT_DIR are needed for the shell script to work properly
INPUT_DIR = "/input"
OUTPUT_DIR = "/output"
RELATIVE_INPUT_DIR = "./input/"
RELATIVE_OUTPUT_DIR = "./output/"
TEST_MODULE_PREFIX = "test_"
TEST_MODULE_FAILING_POSTFIX = "_failing"
FILE_EXTENSION = ".py"

def run(file_name):
    """Runs Pynguin on the given module name to generate unit test cases, and returns the results output

    Args:
        file_name (string): The name of the module to generate unit test cases for

    Returns:
        string: output results from running the PynguinAPI script
    """
    module_name = utils.get_file_name_without_extension(file_name)

    return run_unix(module_name) if utils.is_unix() else run_windows(module_name)

def run_unix(module_name):

    args = shlex.split(f"{TARGET_SCRIPT} {INPUT_DIR} {OUTPUT_DIR} {module_name}")
    return subprocess.check_output(args).decode('utf-8')

    # alternative:
    # stream = os.popen(f"{target_script} {input_dir} {output_dir} {module_name}")
    # output = stream.read()
    # return output

def run_windows(module_name):
    pass


def count_test_cases(file_path):

    if not os.path.exists(file_path):
        return 0
    counter = 0
    with open(file_path, 'r') as f:
        for line in f:
            if line.startswith("def"):
                counter += 1
    return counter

def count_test_cases_passed(module_name):
    assert not module_name.endswith(FILE_EXTENSION)
    passed_test_cases_file = (
        f'{RELATIVE_OUTPUT_DIR}/{TEST_MODULE_PREFIX}{module_name}{FILE_EXTENSION}'
    )
    return count_test_cases(passed_test_cases_file)


def count_test_cases_failed(module_name):
    assert not module_name.endswith(FILE_EXTENSION)
    failed_test_cases_file = (
        f'{RELATIVE_OUTPUT_DIR}/{TEST_MODULE_PREFIX}{module_name}{TEST_MODULE_FAILING_POSTFIX}{FILE_EXTENSION}'
    )
    return count_test_cases(failed_test_cases_file)

def count_passed_and_failed_test_cases(file_name):
    module_name = utils.get_file_name_without_extension(file_name)
    return {
        "passed_test_cases": count_test_cases_passed(module_name),
        "failed_test_cases": count_test_cases_failed(module_name)
    }



def generate_data_set(file_name, number_of_test_cases):

    global data_set
    file_path = f"./integration/{utils.get_file_name_without_extension(file_name)}.xlsx"
    remaining_test_cases = number_of_test_cases

    while (remaining_test_cases > 0):

        # generate test cases
        run(file_name)

        # parse the test cases file and retrieve the input parameters for each function
        get_functions_parameters(f'{RELATIVE_OUTPUT_DIR}{TEST_MODULE_PREFIX}{file_name}')

        # update remaining amount of test cases needed for the next run
        remaining_test_cases -= count_test_cases_passed(utils.get_file_name_without_extension(file_name))

    pprint.pprint(data_set)

    # write data_set values to excel file
    workbook = load_workbook(filename=file_path)
    reach_intersection_sheet = workbook["ReachIntersection"]
    speed_at_intersection_sheet = workbook["SpeedAtIntersection"]
    set_final_speed_sheet = workbook["SetFinalSpeed"]

    # initially assume the first row is the non empty one
    non_empty_reached_intersection_index = 1
    non_empty_speed_at_intersection_index = 1
    non_empty_final_speed_index = 1
    for row in reach_intersection_sheet.iter_rows(values_only=True):
        if row[0] is None:
            break
        non_empty_reached_intersection_index += 1
    for row in speed_at_intersection_sheet.iter_rows(values_only=True):
        if row[0] is None:
            break
        non_empty_speed_at_intersection_index += 1
    for row in set_final_speed_sheet.iter_rows(values_only=True):
        if row[0] is None:
            break
        non_empty_final_speed_index += 1

    for function1_params in data_set[function1]:
        reach_intersection_sheet[f"A{non_empty_reached_intersection_index}"] = function1_params[0]
        reach_intersection_sheet[f"B{non_empty_reached_intersection_index}"] = function1_params[1]
        reach_intersection_sheet[f"C{non_empty_reached_intersection_index}"] = function1_params[2]
        reach_intersection_sheet[f"D{non_empty_reached_intersection_index}"] = function1_params[3]
        non_empty_reached_intersection_index += 1

    for function2_params in data_set[function2]:
        speed_at_intersection_sheet[f"A{non_empty_speed_at_intersection_index}"] = function2_params[0]
        speed_at_intersection_sheet[f"B{non_empty_speed_at_intersection_index}"] = function2_params[1]
        speed_at_intersection_sheet[f"C{non_empty_speed_at_intersection_index}"] = function2_params[2]
        non_empty_speed_at_intersection_index += 1

    for function3_params in data_set[function3]:
        set_final_speed_sheet[f"A{non_empty_final_speed_index}"] = function3_params[0]
        set_final_speed_sheet[f"B{non_empty_final_speed_index}"] = function3_params[1]
        set_final_speed_sheet[f"C{non_empty_final_speed_index}"] = function3_params[2]
        set_final_speed_sheet[f"D{non_empty_final_speed_index}"] = function3_params[3]
        non_empty_final_speed_index += 1

    workbook.save(filename=file_path)

    # clear the in-memory the data_set
    data_set = {
        function1: [],
        function2: [],
        function3: [],
        }

generate_data_set("speed_at_intersection.py", 10)